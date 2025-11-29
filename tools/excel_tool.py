# tools/excel_tool.py
import os
import tempfile
from datetime import datetime
from google.adk.tools import tool
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

@tool
def generate_excel_itinerary(itinerary: list, filename_prefix: str = "itinerary") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Itinerary"

    headers = ["Day", "Time", "Activity", "Location", "Google Maps URL", "Booking Link", "Estimated Cost", "Notes"]
    ws.append(headers)

    for row in itinerary:
        ws.append([
            row.get("day", ""),
            row.get("time", ""),
            row.get("activity", ""),
            row.get("location", ""),
            row.get("maps_url", ""),
            row.get("booking_url", ""),
            row.get("cost", ""),
            row.get("notes", "")
        ])

    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = max_length + 3

    t = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{filename_prefix}_{t}.xlsx"
    tmp_dir = tempfile.gettempdir()
    path = os.path.join(tmp_dir, fname)
    wb.save(path)
    return os.path.abspath(path)
